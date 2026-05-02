"""
User Routes Consolidated
Organized by modules: Core Pages, Asset Master, Spares Requirements, and Breakdown Reports.
"""

import io
import csv
import traceback
from datetime import datetime, timedelta, timezone

from flask import Blueprint, render_template, current_app, jsonify, request, session
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

from services import require_role

# Blueprint (registered under `/user` in app.py)
user_bp = Blueprint("user", __name__)

# ==========================================
# CONSTANTS & TIMEZONE HELPERS
# ==========================================
IST = timezone(timedelta(hours=5, minutes=30))
UTC = timezone.utc

def ist_to_utc(val):
    if not val:
        return None
    try:
        if isinstance(val, datetime):
            dt = val
        elif isinstance(val, str) and "T" in val:
            dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        elif isinstance(val, str):
            dt = datetime.strptime(val, "%d/%m/%Y %I:%M %p")
        else:
            return None

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=IST)
            
        return dt.astimezone(UTC).isoformat()
    except Exception as e:
        raise ValueError(f"Invalid datetime value: {val}") from e

def json_safe(val):
    if isinstance(val, datetime):
        return val.isoformat()
    return val

def utc_to_ist(dt):
    if not dt:
        return None
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(IST)

def _to_iso(val):
    if not val:
        return None
    try:
        if isinstance(val, str):
            return val
        return val.isoformat()
    except Exception:
        try:
            return str(val)
        except Exception:
            return None

def _format_dt_to_ist_string(val):
    if not val:
        return None
    try:
        dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        return dt.astimezone(IST).strftime("%d/%m/%Y %I:%M %p")
    except Exception:
        return None

def _safe_fromiso(val):
    if not val:
        return None
    try:
        if isinstance(val, str):
            dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        elif isinstance(val, datetime):
            dt = val
        else:
            return None

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=IST)
        else:
            dt = dt.astimezone(IST)
            
        return dt
    except Exception:
        raise ValueError("Invalid ISO datetime")


# ==========================================
# MODULE 1: CORE / BASE PAGES
# ==========================================

@user_bp.route("/dashboard")
@require_role("user")
def user_dashboard():
    return render_template("user_dashboard.html")

@user_bp.route("/profile")
@require_role("user")
def user_profile():
    user_email = session.get("user")
    user_role = session.get("role")
    return render_template("user_profile.html", user_email=user_email, user_role=user_role)

@user_bp.route("/<module_name>")
@require_role("user")
def user_module_page(module_name):
    accesses = session.get("accesses", [])
    prefixed = f"user_{module_name}"
    
    if prefixed not in accesses:
        current_app.logger.info(f"Access denied to {prefixed}; redirecting to dashboard")
        return render_template("user_dashboard.html")
        
    try:
        return render_template(f"user_{module_name}.html")
    except Exception as e:
        current_app.logger.warning(f"Missing template user_{module_name}: {e}")
        return render_template("user_asset_master.html")


# ==========================================
# MODULE 2: ASSET MASTER 
# ==========================================

@user_bp.route("/get_assets")
@require_role("user")
def user_get_assets():
    """Return the FULL list of active assets for the User Asset Master table."""
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")
            
        res = supabase_admin.table("asset_master").select("*").order("id").execute()
        return jsonify(res.data or []), 200
        
    except Exception as e:
        current_app.logger.error(f"user_get_assets error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@user_bp.route("/get_dehired_assets")
@require_role("user")
def user_get_dehired_assets():
    """Return the list of de-hired assets for the User view."""
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")
            
        # FIXED: Use "de_hired_assets" to match admin_routes.py perfectly
        # Removed server-side ordering to prevent crashes on missing dates
        res = supabase_admin.table("de_hired_assets").select("*").execute()
        return jsonify(res.data or []), 200
        
    except Exception as e:
        current_app.logger.error(f"user_get_dehired_assets error: {e}")
        return jsonify({"error": str(e)}), 500

@user_bp.route("/dropdown_config", methods=["GET"])
@require_role("user")
def user_get_dropdown_config():
    """Fetch Master Dropdown configs to populate selectors."""
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        if request.args.get('refresh') != '1' and session.get('dropdown_config'):
            return jsonify(session.get('dropdown_config')), 200

        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")
        
        retries = int(current_app.config.get('SUPABASE_HTTP_RETRIES', 3))
        backoff = 1.0
        result = None
        last_exc = None
        
        from httpx import ConnectTimeout
        
        for attempt in range(1, retries + 1):
            try:
                result = supabase_admin.table("dropdown_config").select("*").execute()
                break
            except ConnectTimeout as ct:
                last_exc = ct
                current_app.logger.warning(f"dropdown_config handshake timeout (attempt {attempt}): {ct}")
            except Exception as ex:
                last_exc = ex
                current_app.logger.warning(f"dropdown_config error (attempt {attempt}): {ex}")
                
            if attempt < retries:
                import time
                time.sleep(backoff)
                backoff *= 2

        if result is None:
            raise last_exc or RuntimeError("Failed to fetch dropdown_config")
        
        data = sorted(result.data or [], key=lambda x: (x.get("list_name", ""), x.get("value", "")))
        grouped = {}
        
        for row in data:
            name = row.get("list_name") or "default"
            if name not in grouped:
                grouped[name] = []
            grouped[name].append(row.get("value"))

        try:
            session['dropdown_config'] = grouped
        except Exception:
            current_app.logger.warning("Could not cache dropdown_config in session")

        return jsonify(grouped), 200
        
    except Exception as e:
        current_app.logger.error(f"user_dropdown_config GET error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@user_bp.route("/assets_autocomplete")
@require_role("user")
def assets_autocomplete():
    """Autocomplete for assets used across user workflows (breakdowns/spares)."""
    supabase_admin = current_app.config.get("supabase_admin")
    q = (request.args.get("q") or "").strip()
    
    try:
        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")
        if not q:
            return jsonify([]), 200
        
        res = supabase_admin.table("asset_master").select(
            "asset_code, asset_description, owner, package, location"
        ).ilike("asset_code", f"%{q}%").limit(50).execute()
        
        out = []
        for r in (res.data or []):
            out.append({
                "asset_code": r.get("asset_code"),
                "asset_description": r.get("asset_description"),
                "owner": r.get("owner"),
                "package": r.get("package"),
                "location": r.get("location"),
            })
            
        return jsonify(out), 200
        
    except Exception as e:
        current_app.logger.error(f"assets_autocomplete error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@user_bp.route("/update_asset/<int:asset_id>", methods=["POST"])
@require_role("user")
def user_update_asset(asset_id):
    """
    NEW: User-specific asset update route. 
    Strictly whitelists ONLY the permitted fields defined in the User Asset Master UI.
    """
    supabase_admin = current_app.config.get("supabase_admin")
    data = request.get_json() or {}
    
    try:
        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")

        # STRICT Whitelist of fields a User is allowed to edit
        allowed_fields = {
            "package", "location", "asset_incharge", "asset_user",
            "activity", "activity_works", "supervisor_owner_name",
            "supervisor_owner_phone", "operator1", "operator1_phone",
            "operator1_shift", "operator2", "operator2_phone", "operator2_shift"
        }

        # Filter payload safely based on whitelist
        payload = {}
        for key in data.keys():
            if key in allowed_fields:
                payload[key] = data.get(key)

        # Inject audit trails
        payload["last_updated_at"] = datetime.now(timezone.utc).isoformat()
        payload["last_updated_by"] = session.get("name", session.get("user"))

        # Perform the update
        supabase_admin.table("asset_master").update(payload).eq("id", asset_id).execute()
        return jsonify({"success": True}), 200
        
    except Exception as e:
        current_app.logger.error(f"user_update_asset error: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500


# ==========================================
# MODULE 3: SPARES REQUIREMENTS
# ==========================================

@user_bp.route("/spares_requirements")
@require_role("user")
def user_spares_page():
    user = {
        "username": session.get("user"), 
        "get_full_name": session.get("name", session.get("user"))
    }
    return render_template("user_spares_requirements.html", user=user)

@user_bp.route("/get_spares")
@require_role("user")
def user_get_spares():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")

        try:
            res = supabase_admin.table("spares_requirements").select("*").order("created_at", desc=True).execute()
            rows = res.data or []
        except Exception as e:
            res = supabase_admin.table("spares_requirements").select("*").execute()
            rows = res.data or []
            try:
                rows = sorted(rows, key=lambda x: x.get("id", 0), reverse=True)
            except Exception:
                pass

        out = []
        for r in rows:
            created_raw = r.get("created_at")
            updated_raw = r.get("last_updated_at") or r.get("status_updated_at")
            expected_raw = r.get("expected_date")
            
            asset_display = r.get("asset_code") or ""
            if r.get("asset_description"):
                asset_display += " - " + r.get("asset_description")

            out.append({
                "id": r.get("id"),
                "ref_no": r.get("ref_no"),
                "status": r.get("status"),
                "priority": r.get("priority"),
                "for_type": r.get("for_type"),
                "asset_code": r.get("asset_code"),
                "asset_description": r.get("asset_description"),
                "asset_display": asset_display,
                "spares_req": r.get("spares_req"),
                "qty_required": r.get("qty_required"),
                "qty_available": r.get("qty_available"),
                "required_by": r.get("required_by"),
                "requisition": r.get("requisition") or r.get("created_by") or r.get("requested_by"),
                "actioner": r.get("actioner"),
                "current_status": r.get("current_status"),
                "dc_required": r.get("dc_required", False),
                "dc_number": r.get("dc_number"),
                "expected_date": expected_raw,
                "closed": r.get("closed", False),
                "created_at_iso": _to_iso(created_raw),
                "last_updated_at_iso": _to_iso(updated_raw),
                "expected_date_iso": _to_iso(expected_raw),
                "created_at": _format_dt_to_ist_string(created_raw) or (r.get("created_at") or ""),
                "last_updated_at": _format_dt_to_ist_string(updated_raw) or (r.get("last_updated_at") or ""),
            })
            
        return jsonify(out), 200
        
    except Exception as e:
        current_app.logger.error(f"user_get_spares error: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@user_bp.route("/create_spare", methods=["POST"])
@require_role("user")
def user_create_spare():
    supabase_admin = current_app.config.get("supabase_admin")
    data = request.get_json() or {}
    
    try:
        if not data.get("ref_no") or not data.get("spares_req"):
            return jsonify({"success": False, "error": "ref_no and spares_req are required"}), 400
            
        if not supabase_admin:
            raise RuntimeError("supabase_admin not configured")

        base = {
            "ref_no": data.get("ref_no"),
            "status": data.get("status") or "Active",
            "priority": data.get("priority"),
            "for_type": data.get("for_type"),
            "asset_code": data.get("asset_code"),
            "asset_description": data.get("asset_description"),
            "spares_req": data.get("spares_req"),
            "qty_required": float(data.get("qty_required") or 0),
            "qty_available": float(data.get("qty_available") or 0),
            "required_by": data.get("required_by") or None,
            "requisition": data.get("requisition") or session.get("name", session.get("user")),
            "actioner": data.get("actioner") or session.get("name", session.get("user")),
            "current_status": data.get("current_status") or None,
            "dc_required": bool(data.get("dc_required")) if "dc_required" in data else False,
            "dc_number": data.get("dc_number"),
            "expected_date": data.get("expected_date") if data.get("expected_date") else None,
            "closed": bool(data.get("closed")) if "closed" in data else False,
            "created_at": datetime.now(UTC).isoformat(),
            "updated_at": datetime.now(UTC).isoformat(),
            "last_updated_by": session.get("name", session.get("user")),
        }

        try:
            supabase_admin.table("spares_requirements").insert(base).execute()
            return jsonify({"success": True}), 201
            
        except Exception as ex_insert:
            current_app.logger.warning(f"create_spare initial insert failed: {ex_insert} — retrying without optional fields")
            
            for optional in ("expected_date", "last_updated_at", "last_updated_by", "created_at"):
                base.pop(optional, None)
                
            try:
                supabase_admin.table("spares_requirements").insert(base).execute()
                return jsonify({"success": True, "warning": "insert retried without some optional fields"}), 201
            except Exception as ex_retry:
                return jsonify({"success": False, "error": str(ex_retry)}), 500

    except ValueError:
        return jsonify({"success": False, "error": "Invalid numeric value"}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@user_bp.route("/get_spares_next_ref")
@require_role("user")
def user_get_spares_next_ref():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        res = supabase_admin.table("spares_requirements").select("ref_no").order("id", desc=True).limit(1).execute()
        if res.data and len(res.data) > 0:
            next_num = str(int(res.data[0].get("ref_no")) + 1).zfill(4)
        else:
            next_num = "0001"
        return jsonify({"next_ref": next_num}), 200
    except Exception:
        return jsonify({"next_ref": "0001"}), 200

@user_bp.route("/get_spares_counts")
@require_role("user")
def user_get_spares_counts():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        active_q = supabase_admin.table("spares_requirements").select("id", count="exact").eq("status", "Active").execute()
        pending_q = supabase_admin.table("spares_requirements").select("id", count="exact").eq("status", "Pending").execute()
        total_q = supabase_admin.table("spares_requirements").select("id", count="exact").execute()

        def _get_count(q):
            if hasattr(q, "count") and q.count is not None:
                return q.count
            return len(q.data or [])

        counts = {
            "active": _get_count(active_q),
            "pending": _get_count(pending_q),
            "total": _get_count(total_q)
        }
        
        return jsonify({
            "counts": counts, 
            "updated_at": datetime.now(UTC).isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@user_bp.route("/update_spare/<int:spare_id>", methods=["POST"])
@require_role("user")
def user_update_spare(spare_id):
    supabase_admin = current_app.config.get("supabase_admin")
    data = request.get_json() or {}
    
    try:
        allowed_fields = {
            "current_status", "dc_required", "dc_number", "priority", 
            "status", "remarks", "expected_date", "qty_required", "qty_available"
        }
        
        payload = {}
        for key in data.keys():
            if key in allowed_fields:
                payload[key] = data.get(key)
                
        payload["last_updated_at"] = datetime.now(UTC).isoformat()
        payload["last_updated_by"] = session.get("name", session.get("user"))

        supabase_admin.table("spares_requirements").update(payload).eq("id", spare_id).execute()
        return jsonify({"success": True}), 200
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@user_bp.route("/close_spare/<int:spare_id>", methods=["POST"])
@require_role("user")
def user_close_spare(spare_id):
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        payload = {
            "status": "Closed", 
            "closed": True, 
            "last_updated_at": datetime.now(UTC).isoformat(), 
            "last_updated_by": session.get("name", session.get("user"))
        }
        supabase_admin.table("spares_requirements").update(payload).eq("id", spare_id).execute()
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ==========================================
# MODULE 4: BREAKDOWN REPORTS
# ==========================================

@user_bp.route("/breakdown_report")
@require_role("user")
def user_breakdown_report_page():
    supabase_admin = current_app.config.get("supabase_admin")
    assets = []
    try:
        if supabase_admin:
            res = supabase_admin.table("asset_master").select("asset_code, asset_description, package, owner, location").execute()
            assets = res.data or []
    except Exception as e:
        current_app.logger.warning(f"Could not load asset_master for breakdown page: {e}")
        
    return render_template("user_breakdown_report.html", asset_master=assets)

@user_bp.route("/breakdown_reports", methods=["GET"])
@require_role("user")
def get_breakdown_reports():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        res = supabase_admin.table("breakdown_reports").select("*").order("id", desc=True).execute()
        rows = res.data or []
        now = datetime.now(IST)
        output = []

        for r in rows:
            start_raw = r.get("breakdown_start")
            end_raw = r.get("breakdown_end")
            
            start_dt = _safe_fromiso(start_raw)
            end_dt = _safe_fromiso(end_raw)
            downtime = None

            if start_dt:
                if end_dt:
                    delta = end_dt - start_dt
                else:
                    delta = now - start_dt
                downtime = round(delta.total_seconds() / 3600, 2)
            
            row_out = dict(r)
            row_out["breakdown_start"] = _format_dt_to_ist_string(start_raw)
            row_out["breakdown_end"]   = _format_dt_to_ist_string(end_raw)
            row_out["downtime_hrs"] = downtime
            row_out["created_at"] = _format_dt_to_ist_string(r.get("created_at"))
            output.append(row_out)

        return jsonify(output), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route("/breakdown_reports", methods=["POST"])
@require_role("user")
def create_breakdown_report():
    supabase_admin = current_app.config.get("supabase_admin")
    data = request.get_json() or {}

    missing = []
    for f in ["asset_code", "breakdown_start"]:
        if not data.get(f):
            missing.append(f)
            
    if missing:
        return jsonify({"success": False, "error": f"Missing required fields: {', '.join(missing)}"}), 400

    if data.get("own_hire"):
        data["own_hire"] = data["own_hire"].strip().upper()

    allowed_fields = {
        "asset_code", "asset_description", "asset_package", "own_hire", 
        "agency", "breakdown_start", "breakdown_end", "breakdown_type", 
        "breakdown_description", "root_cause", "responsible_person", 
        "expected_commissioned_at", "eip_commissioned_at", "downtime_hrs", 
        "reported_by", "remarks", "location"
    }
    
    payload = {}
    for k in allowed_fields:
        if k in data:
            payload[k] = data.get(k)

    for f in ("breakdown_start", "breakdown_end", "expected_commissioned_at", "eip_commissioned_at"):
        if payload.get(f):
            payload[f] = ist_to_utc(payload[f])

    now_utc = datetime.now(UTC).isoformat()
    
    payload.update({
        "created_at": now_utc, 
        "updated_at": now_utc, 
        "status": "Active", 
        "reported_by": payload.get("reported_by") or session.get("name", session.get("user")), 
        "last_updated_by": session.get("name", session.get("user"))
    })

    if not payload.get("agency") and payload.get("asset_code"):
        am = supabase_admin.table("asset_master").select("agency").eq("asset_code", payload["asset_code"]).single().execute()
        if am.data:
            payload["agency"] = am.data.get("agency")

    # Safe json conversion
    safe_payload = {}
    for k, v in payload.items():
        safe_payload[k] = json_safe(v)

    try:
        supabase_admin.table("breakdown_reports").insert(safe_payload).execute()
        return jsonify({"success": True}), 201
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@user_bp.route("/breakdown_reports/<int:report_id>", methods=["PUT"])
@require_role("user")
def update_breakdown_report(report_id):
    supabase_admin = current_app.config.get("supabase_admin")
    data = request.get_json() or {}

    try:
        existing = supabase_admin.table("breakdown_reports").select("*").eq("id", report_id).single().execute()
        row = existing.data or {}

        if row.get("status") == "Closed" and not (data.get("breakdown_end") and data.get("eip_commissioned_at")):
            return jsonify({"error": "Closed breakdowns cannot be edited"}), 400

        allowed_fields = {
            "location", "breakdown_end", "breakdown_type", "root_cause", 
            "breakdown_description", "responsible_person", "expected_commissioned_at", 
            "eip_commissioned_at", "remarks"
        }
        
        payload = {}
        for k in allowed_fields:
            if k in data:
                payload[k] = data.get(k)

        for f in ("breakdown_end", "expected_commissioned_at", "eip_commissioned_at"):
            if payload.get(f):
                payload[f] = ist_to_utc(payload[f])

        if payload.get("breakdown_end") and payload.get("eip_commissioned_at"):
            now = datetime.now(UTC)
            start_dt = _safe_fromiso(row.get("breakdown_start")).astimezone(UTC)
            end_dt = payload.get("breakdown_end")
            eip_dt = payload.get("eip_commissioned_at")

            if end_dt <= start_dt:
                return jsonify({"error": "Breakdown End Date must be AFTER Breakdown Start."}), 400
            if eip_dt <= start_dt:
                return jsonify({"error": "EIP Commission Date must be AFTER Breakdown Start."}), 400
            if end_dt > now or eip_dt > now:
                return jsonify({"error": "Dates cannot be in the future."}), 400

            payload.update({
                "status": "Closed", 
                "current_status": "Breakdown Closed", 
                "closed_by": session.get("name") or session.get("user"), 
                "closed_at": now.isoformat()
            })

        payload.update({
            "updated_by": session.get("name", session.get("user")), 
            "updated_at": datetime.now(UTC).isoformat()
        })
        
        supabase_admin.table("breakdown_reports").update(payload).eq("id", report_id).execute()
        return jsonify({"success": True}), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route("/breakdown_summary")
@require_role("user")
def get_breakdown_summary():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        res = supabase_admin.table("breakdown_reports").select("*").execute()
        rows = res.data or []
        now = datetime.now(IST)
        packages = {}
        totals = {"ACTIVE_COUNT": 0, "ACTIVE_DOWNTIME_HRS": 0.0, "TOTAL_COUNT": 0}
        ui_packages = {}
        ui_totals = {"OWN": 0, "HIRE": 0}

        for r in rows:
            pkg = r.get("asset_package") or "Unknown"
            totals["TOTAL_COUNT"] += 1
            
            if pkg not in packages:
                packages[pkg] = {"ACTIVE_COUNT": 0, "ACTIVE_DOWNTIME_HRS": 0.0, "TOTAL_COUNT": 0}
                
            packages[pkg]["TOTAL_COUNT"] += 1

            status_str = str(r.get("status") or "").strip().lower()
            current_status_str = str(r.get("current_status") or "").strip().lower()
            
            is_closed = (status_str == "closed" or "closed" in current_status_str)

            if not is_closed:
                packages[pkg]["ACTIVE_COUNT"] += 1
                totals["ACTIVE_COUNT"] += 1
                
                start_dt = _safe_fromiso(r.get("breakdown_start"))
                end_dt = _safe_fromiso(r.get("breakdown_end"))
                
                if start_dt:
                    if end_dt:
                        delta = end_dt - start_dt
                    else:
                        delta = now - start_dt
                        
                    hrs = round(delta.total_seconds() / 3600, 2)
                    packages[pkg]["ACTIVE_DOWNTIME_HRS"] += hrs
                    totals["ACTIVE_DOWNTIME_HRS"] += hrs

            oh = (r.get("own_hire") or "").upper()
            if pkg not in ui_packages:
                ui_packages[pkg] = {"OWN": 0, "HIRE": 0}
                
            if oh in ("OWN", "HIRE"):
                ui_packages[pkg][oh] += 1
                ui_totals[oh] += 1

        ui_totals["ALL"] = ui_totals["OWN"] + ui_totals["HIRE"]
        
        return jsonify({
            "packages": packages, 
            "totals": totals, 
            "own_hire": {
                "packages": ui_packages, 
                "totals": ui_totals
            }
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route("/breakdown_dashboard")
@require_role("user")
def get_breakdown_dashboard():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        res = supabase_admin.table("breakdown_reports").select("*").execute()
        rows = res.data or []
        now = datetime.now(IST)

        total_count = 0
        active_count = 0
        closed_count = 0
        active_delay_sum = 0.0
        closed_repair_sum = 0.0
        closed_repair_count = 0
        
        packages = {}
        ageing = {"0_24": 0, "24_48": 0, "48_plus": 0}
        own_hire = {"OWN": {"count": 0, "repair_sum": 0.0}, "HIRE": {"count": 0, "repair_sum": 0.0}}

        for r in rows:
            total_count += 1
            pkg = r.get("asset_package") or "Unknown"
            oh = (r.get("own_hire") or "").upper()
            
            status_str = str(r.get("status") or "").strip().lower()
            current_status_str = str(r.get("current_status") or "").strip().lower()
            is_closed = (status_str == "closed" or "closed" in current_status_str)
            
            if pkg not in packages:
                packages[pkg] = {"ACTIVE": 0, "ACTIVE_DELAY_SUM": 0.0, "CLOSED": 0, "CLOSED_REPAIR_SUM": 0.0}

            start_dt = _safe_fromiso(r.get("breakdown_start"))
            end_dt = _safe_fromiso(r.get("breakdown_end"))
            
            if not start_dt:
                continue

            if is_closed:
                closed_count += 1
                packages[pkg]["CLOSED"] += 1
                if end_dt:
                    hrs = round((end_dt - start_dt).total_seconds() / 3600, 2)
                    closed_repair_sum += hrs
                    closed_repair_count += 1
                    packages[pkg]["CLOSED_REPAIR_SUM"] += hrs
                    if oh in own_hire:
                        own_hire[oh]["count"] += 1
                        own_hire[oh]["repair_sum"] += hrs
            else:
                active_count += 1
                packages[pkg]["ACTIVE"] += 1
                hrs = round((now - start_dt).total_seconds() / 3600, 2)
                active_delay_sum += hrs
                packages[pkg]["ACTIVE_DELAY_SUM"] += hrs
                
                if hrs <= 24:
                    ageing["0_24"] += 1
                elif hrs <= 48:
                    ageing["24_48"] += 1
                else:
                    ageing["48_plus"] += 1

        avg_active_delay = 0.0
        if active_count > 0:
            avg_active_delay = active_delay_sum / active_count
            
        avg_repair_time = 0.0
        if closed_repair_count > 0:
            avg_repair_time = closed_repair_sum / closed_repair_count

        own_hire_result = {}
        for k, v in own_hire.items():
            if v["count"] > 0:
                own_hire_result[k] = v["repair_sum"] / v["count"]
            else:
                own_hire_result[k] = 0.0
                
        package_result = {}
        for pkg, p in packages.items():
            pkg_avg_delay = 0.0
            if p["ACTIVE"] > 0:
                pkg_avg_delay = p["ACTIVE_DELAY_SUM"] / p["ACTIVE"]
                
            pkg_avg_repair = 0.0
            if p["CLOSED"] > 0:
                pkg_avg_repair = p["CLOSED_REPAIR_SUM"] / p["CLOSED"]
                
            package_result[pkg] = {
                "active": p["ACTIVE"], 
                "avg_active_delay": pkg_avg_delay, 
                "avg_repair_time": pkg_avg_repair
            }

        return jsonify({
            "counts": {"total": total_count, "active": active_count, "closed": closed_count}, 
            "kpi": {"avg_active_delay": round(avg_active_delay, 2), "avg_repair_time": round(avg_repair_time, 2)}, 
            "packages": package_result, 
            "own_hire": own_hire_result, 
            "ageing": ageing, 
            "rows": rows
        }), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route("/breakdown_reports/export")
@require_role("user")
def export_breakdown_reports():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        res = supabase_admin.table("breakdown_reports").select("*").order("id", desc=True).execute()
        rows = res.data or []
        now = datetime.now(IST)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "id", "asset_code", "asset_description", "asset_package", "own_hire", 
            "agency", "location", "breakdown_start", "breakdown_end", "downtime_hrs", 
            "breakdown_type", "root_cause", "breakdown_description", "status", 
            "current_status", "responsible_person", "expected_commissioned_at", 
            "eip_commissioned_at", "reported_by", "created_by", "updated_by", 
            "created_at", "remarks"
        ])

        for r in rows:
            start_raw = r.get("breakdown_start")
            end_raw = r.get("breakdown_end")
            start_dt = _safe_fromiso(start_raw)
            end_dt = _safe_fromiso(end_raw)
            downtime = ""
            
            if start_dt:
                if end_dt:
                    downtime = round((end_dt - start_dt).total_seconds() / 3600, 2)
                else:
                    downtime = round((now - start_dt).total_seconds() / 3600, 2)
            
            writer.writerow([
                r.get("id"), r.get("asset_code"), r.get("asset_description"), 
                r.get("asset_package"), r.get("own_hire"), r.get("agency"), 
                r.get("location"), _to_iso(start_raw), _to_iso(end_raw), downtime, 
                r.get("breakdown_type"), r.get("root_cause"), r.get("breakdown_description"), 
                r.get("status"), r.get("current_status"), r.get("responsible_person"), 
                _to_iso(r.get("expected_commissioned_at")), _to_iso(r.get("eip_commissioned_at")), 
                r.get("reported_by"), r.get("created_by"), r.get("updated_by"), 
                _to_iso(r.get("created_at")), r.get("remarks")
            ])

        output.seek(0)
        return current_app.response_class(
            output.getvalue().encode("utf-8"), 
            mimetype="text/csv", 
            headers={"Content-Disposition": "attachment;filename=breakdown_reports.csv"}
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route("/breakdown_reports/export_xlsx")
@require_role("user")
def export_breakdown_reports_xlsx():
    supabase_admin = current_app.config.get("supabase_admin")
    try:
        res = supabase_admin.table("breakdown_reports").select("*").order("id", desc=True).execute()
        rows = res.data or []
        now = datetime.now(IST)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Breakdown Reports"
        
        header = [
            "id", "asset_code", "asset_description", "asset_package", "own_hire", 
            "agency", "location", "breakdown_start", "breakdown_end", "downtime_hrs", 
            "breakdown_type", "root_cause", "breakdown_description", "status", 
            "current_status", "responsible_person", "expected_commissioned_at", 
            "eip_commissioned_at", "reported_by", "created_by", "updated_by", 
            "created_at", "remarks"
        ]

        for c, h in enumerate(header, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        border = Border(
            left=Side(border_style="thin", color="000000"), 
            right=Side(border_style="thin", color="000000"), 
            top=Side(border_style="thin", color="000000"), 
            bottom=Side(border_style="thin", color="000000")
        )

        for r_idx, r in enumerate(rows, start=2):
            start_raw = r.get("breakdown_start")
            end_raw = r.get("breakdown_end")
            start_dt = _safe_fromiso(start_raw)
            end_dt = _safe_fromiso(end_raw)
            downtime = ""
            
            if start_dt:
                if end_dt:
                    downtime = round((end_dt - start_dt).total_seconds() / 3600, 2)
                else:
                    downtime = round((now - start_dt).total_seconds() / 3600, 2)
            
            vals = [
                r.get("id"), r.get("asset_code"), r.get("asset_description"), 
                r.get("asset_package"), r.get("own_hire"), r.get("agency"), 
                r.get("location"), _to_iso(start_raw), _to_iso(end_raw), downtime, 
                r.get("breakdown_type"), r.get("root_cause"), r.get("breakdown_description"), 
                r.get("status"), r.get("current_status"), r.get("responsible_person"), 
                _to_iso(r.get("expected_commissioned_at")), _to_iso(r.get("eip_commissioned_at")), 
                r.get("reported_by"), r.get("created_by"), r.get("updated_by"), 
                _to_iso(r.get("created_at")), r.get("remarks")
            ]

            for c_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                cell.border = border

        for i, col in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = (max_length + 2) if max_length > 0 else 12
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = adjusted_width

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        
        return current_app.response_class(
            bio.getvalue(), 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            headers={"Content-Disposition": "attachment;filename=breakdown_reports.xlsx"}
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500